import os
import argparse
from PIL import Image
from tqdm import tqdm
from sklearn.cluster import KMeans
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict

class DiagonalPixelArtConverter:
    def __init__(self, image_path):
        """初始化图片转换器"""
        self.img = Image.open(image_path).convert("RGB")
        self.width, self.height = self.img.size
        self.pixel_data = []
        self.filename = os.path.splitext(os.path.basename(image_path))[0]

    def resize_image(self, max_dimension=100):
        """调整图片尺寸"""
        if self.width > self.height:
            new_width = max_dimension
            new_height = int(self.height * (max_dimension / self.width))
        else:
            new_height = max_dimension
            new_width = int(self.width * (max_dimension / self.height))

        self.img = self.img.resize((new_width, new_height), Image.NEAREST)
        self.width, self.height = self.img.size
        print(f"图片已调整为: {self.width}×{self.height} 像素")

    def reduce_colors(self, n_colors):
        """使用K-means算法减少颜色数量"""
        print(f"正在将颜色减少到 {n_colors} 种...")
        img_array = np.array(self.img)
        h, w, c = img_array.shape
        pixel_samples = img_array.reshape(-1, 3)

        kmeans = KMeans(n_clusters=n_colors, random_state=0).fit(pixel_samples)
        new_colors = kmeans.cluster_centers_.astype(int)
        new_img_array = new_colors[kmeans.labels_].reshape(h, w, c)

        self.img = Image.fromarray(new_img_array.astype("uint8"))
        print(f"颜色已减少到 {n_colors} 种")

    def analyze_pixels(self):
        """分析像素数据并生成对角线编号（从右下角开始）"""
        self.pixel_data = []

        for y in tqdm(range(self.height), desc="处理像素"):
            row = []
            for x in range(self.width):
                r, g, b = self.img.getpixel((x, y))
                hex_color = f"{r:02X}{g:02X}{b:02X}"

                # 对角线编号：从右下角(0)开始，向左上方递增
                diagonal_num = (self.width - 1 - x) + (self.height - 1 - y)

                row.append(
                    {
                        "x": x,
                        "y": y,
                        "diagonal": diagonal_num,
                        "color": (r, g, b),
                        "hex": hex_color,
                    }
                )
            self.pixel_data.append(row)

    def analyze_diagonal_sequences(self):
        """分析每个对角线的连续颜色块序列"""
        diagonal_sequences = {}

        # 获取颜色索引映射
        color_to_index = {}
        for row in self.pixel_data:
            for pixel in row:
                if pixel["color"] not in color_to_index:
                    color_to_index[pixel["color"]] = len(color_to_index) + 1

        # 分析每个对角线
        for diagonal_num in range(self.width + self.height - 1):
            sequence = []

            # 收集该对角线上的所有像素
            diagonal_pixels = []
            for y in range(self.height):
                for x in range(self.width):
                    if (self.width - 1 - x) + (self.height - 1 - y) == diagonal_num:
                        diagonal_pixels.append(self.pixel_data[y][x])

            # 按x坐标排序（从左到右）
            diagonal_pixels.sort(key=lambda p: p["x"])

            # 统计连续颜色块
            if diagonal_pixels:
                current_color = color_to_index[diagonal_pixels[0]["color"]]
                current_count = 1

                for i in range(1, len(diagonal_pixels)):
                    pixel_color = color_to_index[diagonal_pixels[i]["color"]]

                    if pixel_color == current_color:
                        # 相同颜色，计数加1
                        current_count += 1
                    else:
                        # 不同颜色，保存当前块并开始新块
                        sequence.append((current_color, current_count))
                        current_color = pixel_color
                        current_count = 1

                # 保存最后一个颜色块
                sequence.append((current_color, current_count))

            diagonal_sequences[diagonal_num] = sequence

        return diagonal_sequences, color_to_index

    def generate_excel(self, output_dir="output"):
        """生成带颜色和对角线编号的Excel文件"""
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{self.filename}_pixelart.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "像素图"

        # 设置单元格样式
        font = Font(color="FFFFFF", bold=True)  # 白色粗体字体
        alignment = Alignment(horizontal="center", vertical="center")

        # 设置所有单元格为统一大小
        for col in range(1, self.width + 10):  # 预留足够列数给对角线构成
            ws.column_dimensions[get_column_letter(col)].width = 5

        for row in range(1, self.width + self.height + 5):  # 预留足够行数给对角线统计
            ws.row_dimensions[row].height = 30

        # 填充颜色和对角线编号
        for y in tqdm(range(self.height), desc="生成Excel"):
            for x in range(self.width):
                pixel = self.pixel_data[y][x]
                cell = ws.cell(row=y + 1, column=x + 1)

                # 设置颜色填充
                cell.fill = PatternFill(
                    start_color=pixel["hex"], end_color=pixel["hex"], fill_type="solid"
                )

                # 添加对角线编号
                cell.value = pixel["diagonal"]
                cell.font = font
                cell.alignment = alignment

                # 添加颜色统计表
        self._add_color_sheet(wb)

        # 添加对角线统计表
        self._add_diagonal_sheet(wb)

        wb.save(output_path)
        print(f"\nExcel文件已保存: {os.path.abspath(output_path)}")

    def _add_color_sheet(self, workbook):
        """添加颜色统计表"""
        ws = workbook.create_sheet(title="颜色统计")
        ws.append(["颜色索引", "RGB值", "十六进制", "使用次数", "使用单元格"])

        # 统计颜色使用情况
        color_stats = defaultdict(list)
        for row in self.pixel_data:
            for pixel in row:
                color_stats[pixel["color"]].append((pixel["x"], pixel["y"]))

        # 按使用频率排序
        sorted_colors = sorted(
            color_stats.items(), key=lambda x: len(x[1]), reverse=True
        )

        for idx, (color, pixels) in enumerate(sorted_colors):
            r, g, b = color
            hex_color = f"{r:02X}{g:02X}{b:02X}"

            # 格式化单元格坐标
            cell_refs = ", ".join(
                [f"{get_column_letter(x+1)}{y+1}" for x, y in pixels[:5]]
            )
            if len(pixels) > 5:
                cell_refs += f", ...(共{len(pixels)}个)"

            ws.append([idx + 1, f"({r}, {g}, {b})", hex_color, len(pixels), cell_refs])

            # 设置颜色示例
            cell = ws.cell(row=ws.max_row, column=3)
            cell.fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )

    def _add_diagonal_sheet(self, workbook):
        """添加对角线统计表"""
        ws = workbook.create_sheet(title="对角线统计")

        # 分析对角线颜色序列
        diagonal_sequences, color_to_index = self.analyze_diagonal_sequences()

        # 创建索引到颜色的反向映射
        index_to_color = {v: k for k, v in color_to_index.items()}

        # 设置列标题
        ws.cell(row=1, column=1).value = "对角线索引"
        ws.cell(row=1, column=2).value = "对角线构成"

        # 设置标题样式
        for col in [1, 2]:
            title_cell = ws.cell(row=1, column=col)
            title_cell.font = Font(bold=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 设置单元格大小
        for col in range(1, 20):  # 预留足够列数
            ws.column_dimensions[get_column_letter(col)].width = 5

        for row in range(1, self.width + self.height + 5):
            ws.row_dimensions[row].height = 30

        # 为每个对角线添加序列
        for diagonal_num in range(self.width + self.height - 1):
            row_num = diagonal_num + 2  # 从第2行开始

            # 添加对角线索引号
            diagonal_index_cell = ws.cell(row=row_num, column=1)
            diagonal_index_cell.value = diagonal_num
            diagonal_index_cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )
            diagonal_index_cell.font = Font(bold=True)

            if diagonal_num in diagonal_sequences:
                sequence = diagonal_sequences[diagonal_num]
                col_offset = 0

                for color_index, count in sequence:
                    # 显示数量（带颜色背景）
                    count_cell = ws.cell(row=row_num, column=2 + col_offset)
                    count_cell.value = count
                    count_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                    # 设置背景色
                    rgb_color = index_to_color[color_index]
                    hex_color = (
                        f"{rgb_color[0]:02X}{rgb_color[1]:02X}{rgb_color[2]:02X}"
                    )
                    count_cell.fill = PatternFill(
                        start_color=hex_color, end_color=hex_color, fill_type="solid"
                    )

                    # 设置字体颜色（根据背景色调整）
                    if sum(rgb_color) > 384:  # 浅色背景用黑色字体
                        count_cell.font = Font(color="000000", bold=True)
                    else:  # 深色背景用白色字体
                        count_cell.font = Font(color="FFFFFF", bold=True)

                    col_offset += 1


def main():
    parser = argparse.ArgumentParser(
        description="图片转对角线编号Excel像素图工具（从右下角开始编号）"
    )
    parser.add_argument("image_path", help="输入图片路径")
    parser.add_argument("--size", type=int, default=50, help="最大尺寸（保持宽高比）")
    parser.add_argument("--colors", type=int, help="限制颜色数量（K-means聚类）")

    args = parser.parse_args()

    try:
        converter = DiagonalPixelArtConverter(args.image_path)
        converter.resize_image(args.size)

        if args.colors:
            converter.reduce_colors(args.colors)

        converter.analyze_pixels()
        converter.generate_excel()

        print("\n转换完成！Excel文件包含以下工作表：")
        print("- 像素图：颜色填充和对角线编号（右下角为0，向左上方递增）")
        print("- 颜色统计：颜色索引、RGB值、使用次数和使用位置")
        print("- 对角线统计：每条对角线的颜色构成序列")
    except Exception as e:
        print(f"发生错误: {str(e)}")

if __name__ == "__main__":
    main()